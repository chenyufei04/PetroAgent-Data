from __future__ import annotations

import csv
import io
import re
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
UPLOAD_DIR = PROJECT_ROOT / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MAX_FILE_SIZE = 100 * 1024 * 1024
MAX_PREVIEW_ROWS = 500

app = FastAPI(title="石油工程术语质控本地服务", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "DELETE"],
    allow_headers=["*"],
)


def safe_filename(filename: str | None) -> str:
    name = Path(filename or "uploaded-file").name
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip(" .")
    return name or "uploaded-file"


def available_path(filename: str) -> Path:
    target = UPLOAD_DIR / filename
    if not target.exists():
        return target
    stem, suffix = target.stem, target.suffix
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    return UPLOAD_DIR / f"{stem}-{stamp}{suffix}"


def extract_excel_text(path: Path) -> str | None:
    if path.suffix.lower() != ".xlsx":
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        lines: list[str] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                values = [str(value) if value is not None else "" for value in row]
                if any(values):
                    lines.append(" ".join(values))
        return "\n".join(lines)
    finally:
        workbook.close()


def extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        document = ElementTree.fromstring(archive.read("word/document.xml"))
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    paragraphs: list[str] = []
    for paragraph in document.iter(f"{namespace}p"):
        text = "".join(node.text or "" for node in paragraph.iter(f"{namespace}t"))
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def extract_file_text(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return extract_excel_text(path)
    if suffix == ".docx":
        return extract_docx_text(path)
    if suffix in {".csv", ".tsv", ".txt", ".md", ".json", ".xml"}:
        content = path.read_bytes()
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return content.decode("utf-8", errors="replace")
    return None


def extract_table_data(path: Path) -> list[list[str]] | None:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows: list[list[str]] = []
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(values_only=True):
                    values = [str(value) if value is not None else "" for value in row]
                    if any(values):
                        rows.append(values)
                        if len(rows) >= MAX_PREVIEW_ROWS:
                            return rows
            return rows
        finally:
            workbook.close()
    if suffix in {".csv", ".tsv"}:
        text = extract_file_text(path) or ""
        delimiter = "\t" if suffix == ".tsv" else ","
        return [row for _, row in zip(range(MAX_PREVIEW_ROWS), csv.reader(io.StringIO(text), delimiter=delimiter))]
    return None


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok", "upload_dir": str(UPLOAD_DIR)}


@app.get("/api/quality/uploads")
def list_quality_files() -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for path in sorted(UPLOAD_DIR.iterdir(), key=lambda item: item.stat().st_mtime, reverse=True):
        if not path.is_file() or path.name.startswith("."):
            continue
        parse_error: str | None = None
        try:
            extracted_text = extract_file_text(path)
            table_data = extract_table_data(path)
        except Exception as exc:
            extracted_text = None
            table_data = None
            parse_error = f"文件解析失败：{exc}"
        records.append({
            "filename": path.name,
            "path": str(path),
            "size": path.stat().st_size,
            "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
            "extracted_text": extracted_text,
            "table_data": table_data,
            "parse_error": parse_error,
        })
    return records


@app.post("/api/quality/upload")
async def save_quality_file(file: UploadFile = File(...)) -> dict[str, object]:
    filename = safe_filename(file.filename)
    content = await file.read(MAX_FILE_SIZE + 1)
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="文件不能超过100 MB")
    target = available_path(filename)
    target.write_bytes(content)
    try:
        extracted_text = extract_file_text(target)
        table_data = extract_table_data(target)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"文件解析失败：{exc}") from exc
    return {"filename": target.name, "path": str(target), "size": len(content), "extracted_text": extracted_text, "table_data": table_data}


@app.delete("/api/quality/upload/{filename}")
def delete_quality_file(filename: str) -> dict[str, str]:
    safe_name = safe_filename(filename)
    if safe_name != filename:
        raise HTTPException(status_code=400, detail="文件名不合法")
    target = UPLOAD_DIR / safe_name
    if not target.is_file():
        raise HTTPException(status_code=404, detail="文件不存在或已被删除")
    try:
        target.unlink()
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"删除文件失败：{exc}") from exc
    return {"status": "deleted", "filename": safe_name}
